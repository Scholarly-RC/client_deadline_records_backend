# views.py
import os
from datetime import timedelta

from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.db.models.deletion import RestrictedError
from django.http import FileResponse, Http404, HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from core.actions import (
    create_log,
    create_notifications,
    initiate_task_approval,
    process_task_approval,
)
from core.choices import TaskStatus
from core.models import (
    AppLog,
    Client,
    ClientDocument,
    Notification,
    Task,
    TaskApproval,
    TaskStatusHistory,
    User,
)
from core.pagination import CustomPageNumberPagination
from core.serializers import (
    AppLogSerializer,
    ClientBirthdaySerializer,
    ClientDocumentSerializer,
    ClientSerializer,
    InitiateApprovalSerializer,
    NotificationSerializer,
    ProcessApprovalSerializer,
    TaskApprovalSerializer,
    TaskListSerializer,
    TaskSerializer,
    TaskStatusHistorySerializer,
    UserMiniSerializer,
    UserSerializer,
)
from core.utils import (
    get_admin_users,
    get_notification_recipients,
    get_now_local,
    get_today_local,
)


class IsOwnerOrStaff(permissions.BasePermission):
    """
    Custom permission to only allow owners of an object or staff to edit it.
    """

    def has_object_permission(self, request, view, obj):
        if request.user and request.user.is_admin:
            return True

        # For objects with created_by field
        if hasattr(obj, "created_by"):
            return obj.created_by == request.user

        # For objects with assigned_to field (like Task)
        if hasattr(obj, "assigned_to"):
            return obj.assigned_to == request.user

        return False


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.exclude(is_superuser=True).prefetch_related("logs")
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = UserSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
    ]
    search_fields = ["first_name", "middle_name", "last_name", "email", "username"]

    def perform_create(self, serializer):
        instance = serializer.save()
        create_log(self.request.user, f"Created user: {instance}.")

    def perform_update(self, serializer):
        instance = serializer.save()
        create_log(
            self.request.user,
            f"Updated user: {instance}. Password changed: {'Yes' if 'password' in serializer.validated_data else 'No'}.",
        )

    @action(detail=False, methods=["get"], url_path="get-current-user")
    def get_current_user(self, request):
        user = request.user
        serializer = UserMiniSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="toggle-active-status")
    def toggle_active_status(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        serializer = self.get_serializer(user)
        create_log(
            request.user,
            f"{'Enabled' if user.is_active else 'Disabled'} user: {user.fullname}.",
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="user-choices")
    def get_user_choices(self, request):
        users = self.get_queryset().filter(is_active=True)
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="unread-notification-count")
    def get_unread_notification_count(self, request, pk=None):
        unread_count = self.get_object().notifications.filter(is_read=False).count()
        return Response({"unread_count": unread_count}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="deadlines-tasks")
    def get_user_deadlines_tasks(self, request, pk=None):
        user = self.get_object()

        # Get all tasks assigned to this user
        tasks = user.tasks_assigned_to.all()

        # Apply pagination
        paginator = CustomPageNumberPagination()
        paginated_tasks = paginator.paginate_queryset(tasks, request)

        # Serialize the paginated data
        serializer = TaskListSerializer(paginated_tasks, many=True)

        # Return paginated response
        return paginator.get_paginated_response(serializer.data)


class TaskViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Task records

    Provides CRUD operations for tasks with filtering,
    searching, and ordering capabilities.
    """

    queryset = (
        Task.objects.select_related("assigned_to", "client")
        .prefetch_related(
            "approvals__approver",
            "approvals__next_approver",
            "status_history_records__changed_by",
        )
        .all()
    )
    serializer_class = TaskSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    # Filtering options
    filterset_fields = {
        "client": ["exact"],
        "category": ["exact", "in"],
        "status": ["exact", "in"],
        "priority": ["exact", "in"],
        "assigned_to": ["exact"],
        "engagement_date": ["gte", "lte", "exact"],
        "deadline": ["gte", "lte", "exact"],
        "completion_date": ["gte", "lte", "exact"],
        "tax_category": ["exact", "in"],
        "tax_type": ["exact", "in"],
        "form": ["exact", "in"],
    }

    # Search fields
    search_fields = [
        "description",
        "remarks",
        "steps",
        "requirements",
        "type",
        "needed_data",
        "area",
        "working_paper",
    ]

    # Ordering options
    ordering_fields = [
        "deadline",
        "engagement_date",
        "priority",
        "status",
        "last_update",
        "tax_payable",
        "last_followup",
    ]
    ordering = ["-last_update"]  # Default ordering

    def get_queryset(self):
        """
        Filter queryset based on user permissions.
        Admin users see all records, non-admin users only see records assigned to them.
        """
        queryset = (
            Task.objects.select_related("assigned_to", "client")
            .prefetch_related(
                "approvals__approver",
                "approvals__next_approver",
                "status_history_records__changed_by",
            )
            .all()
        )

        # Return empty queryset for unauthenticated users
        if not self.request.user.is_authenticated:
            return queryset.none()

        if self.request.user.is_admin:
            return queryset

        return queryset.filter(assigned_to=self.request.user)

    def get_serializer_class(self):
        """Use different serializer for list action"""
        if self.action == "list":
            return TaskListSerializer
        return TaskSerializer

    def perform_create(self, serializer):
        """Set last_update timestamp when creating"""
        instance = serializer.save(last_update=get_now_local())
        create_log(self.request.user, f"Created task: {instance}.")

        # Notify assigned user if task is assigned
        if instance.assigned_to and instance.assigned_to != self.request.user:
            create_notifications(
                recipient=instance.assigned_to,
                title="New Task Assigned",
                message=f"A new task '{instance.description}' has been assigned to you.",
                link="/my-deadlines",
            )

    def update(self, request, *args, **kwargs):
        """Override update to track assigned_to changes and general updates"""
        # Get the instance before update
        instance = self.get_object()
        original_assigned_to = instance.assigned_to
        original_data = {
            "description": instance.description,
            "status": instance.status,
            "priority": instance.priority,
            "deadline": instance.deadline,
            "remarks": instance.remarks,
            "period_covered": instance.period_covered,
            "engagement_date": instance.engagement_date,
            "steps": instance.steps,
            "requirements": instance.requirements,
            "type": instance.type,
            "needed_data": instance.needed_data,
            "area": instance.area,
            "tax_category": instance.tax_category,
            "tax_type": instance.tax_type,
            "form": instance.form,
            "working_paper": instance.working_paper,
            "tax_payable": instance.tax_payable,
            "last_followup": instance.last_followup,
        }

        # Perform the update
        response = super().update(request, *args, **kwargs)

        # Get the updated instance
        updated_instance = self.get_object()

        # Check if assigned_to has changed and notify both new and previous assignees
        if (
            original_assigned_to != updated_instance.assigned_to
            and updated_instance.assigned_to
        ):
            # Notify new assignee
            if updated_instance.assigned_to != self.request.user:
                create_notifications(
                    recipient=updated_instance.assigned_to,
                    title=(
                        "Task Reassigned"
                        if original_assigned_to
                        else "New Task Assigned"
                    ),
                    message=f"The task '{updated_instance.description}' has been {'reassigned' if original_assigned_to else 'assigned'} to you.",
                    link="/my-deadlines",
                )

            # Notify previous assignee (only if there was one and it's not the same as the updater)
            if (
                original_assigned_to
                and original_assigned_to != self.request.user
                and original_assigned_to != updated_instance.assigned_to
            ):
                create_notifications(
                    recipient=original_assigned_to,
                    title="Task Reassigned",
                    message=f"The task '{updated_instance.description}' has been reassigned to {updated_instance.assigned_to.fullname}.",
                    link="/my-deadlines",
                )

        # Check if other fields have changed and notify the assigned user
        # (only if assigned_to didn't change or if it changed to someone else)
        current_data = {
            "description": updated_instance.description,
            "status": updated_instance.status,
            "priority": updated_instance.priority,
            "deadline": updated_instance.deadline,
            "remarks": updated_instance.remarks,
            "period_covered": updated_instance.period_covered,
            "engagement_date": updated_instance.engagement_date,
            "steps": updated_instance.steps,
            "requirements": updated_instance.requirements,
            "type": updated_instance.type,
            "needed_data": updated_instance.needed_data,
            "area": updated_instance.area,
            "tax_category": updated_instance.tax_category,
            "tax_type": updated_instance.tax_type,
            "form": updated_instance.form,
            "working_paper": updated_instance.working_paper,
            "tax_payable": updated_instance.tax_payable,
            "last_followup": updated_instance.last_followup,
        }

        # Check if any field has changed
        fields_changed = any(
            original_data[key] != current_data[key] for key in original_data
        )

        # Send notification to assigned user if fields changed and user is not the one making the update
        if (
            fields_changed
            and updated_instance.assigned_to
            and updated_instance.assigned_to != self.request.user
        ):
            # Don't send if this was just a reassignment (already handled above)
            if original_assigned_to == updated_instance.assigned_to:
                create_notifications(
                    recipient=updated_instance.assigned_to,
                    title="Task Updated",
                    message=f"The task '{updated_instance.description}' has been updated.",
                    link="/my-deadlines",
                )

        return response

    def perform_update(self, serializer):
        """Set last_update timestamp when updating"""
        serializer.save(last_update=get_now_local())
        create_log(self.request.user, f"Updated task: {serializer.instance}.")

    @action(detail=False, methods=["get"])
    def overdue(self, request):
        """Get all overdue tasks"""
        today = get_today_local()
        overdue_tasks = self.get_queryset().filter(
            deadline__lt=today,
            status__in=["NOT_YET_STARTED", "IN_PROGRESS"],
        )

        serializer = self.get_serializer(overdue_tasks, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def due_soon(self, request):
        """Get tasks due within the next 7 days"""
        today = get_today_local()
        next_week = today + timedelta(days=7)

        due_soon_tasks = self.get_queryset().filter(
            deadline__gte=today,
            deadline__lte=next_week,
            status__in=["NOT_YET_STARTED", "IN_PROGRESS"],
        )

        serializer = self.get_serializer(due_soon_tasks, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def by_category(self, request):
        """Get tasks grouped by category"""
        category = request.query_params.get("category")
        if not category:
            return Response(
                {"error": "category parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        category_tasks = self.get_queryset().filter(category=category)
        serializer = self.get_serializer(category_tasks, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def by_user(self, request):
        """Get tasks grouped by assigned user"""
        user_id = request.query_params.get("user_id")
        if not user_id:
            return Response(
                {"error": "user_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_tasks = self.get_queryset().filter(assigned_to_id=user_id)
        serializer = self.get_serializer(user_tasks, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def mark_completed(self, request, pk=None):
        """Mark a task as completed"""
        task = self.get_object()

        completion_date = request.data.get("completion_date", get_today_local())
        date_complied = request.data.get("date_complied", get_today_local())
        remarks = request.data.get("remarks", task.remarks)

        task.status = TaskStatus.COMPLETED.value
        task.completion_date = completion_date
        task.date_complied = date_complied
        task.remarks = remarks
        task.last_update = get_now_local()
        task.save()

        # Log the completion
        create_log(
            request.user,
            f"Marked task '{task.description}' as completed. Completion date: {completion_date.strftime('%Y-%m-%d')}",
        )

        serializer = self.get_serializer(task)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def statistics(self, request):
        """Get comprehensive task statistics optimized for dashboard visualization"""
        from datetime import datetime, timedelta

        from django.db.models import Avg, Case, Count, F, IntegerField, Q, Sum, When
        from django.db.models.functions import Extract, TruncMonth

        from core.choices import (
            TaskCategory,
            TaskPriority,
            TaskStatus,
            TaxCaseCategory,
            TypeOfTaxCase,
        )

        queryset = self.get_queryset()

        # Parse and validate date range filter parameters
        start_date_param = request.query_params.get("start_date")
        end_date_param = request.query_params.get("end_date")
        filters_applied = {}
        start_date = None
        end_date = None

        if start_date_param or end_date_param:
            try:
                from datetime import datetime

                # Validate start_date parameter
                if start_date_param:
                    start_date = datetime.strptime(start_date_param, "%Y-%m-%d").date()
                    filters_applied["start_date"] = start_date_param

                # Validate end_date parameter
                if end_date_param:
                    end_date = datetime.strptime(end_date_param, "%Y-%m-%d").date()
                    filters_applied["end_date"] = end_date_param

                # Validate date range logic
                if start_date and end_date and start_date > end_date:
                    return Response(
                        {"error": "start_date cannot be after end_date"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Apply filtering based on last_update field
                if start_date and end_date:
                    # Filter by date range
                    queryset = queryset.filter(
                        last_update__date__range=[start_date, end_date]
                    )
                elif start_date:
                    # Filter from start_date onwards
                    queryset = queryset.filter(last_update__date__gte=start_date)
                elif end_date:
                    # Filter up to end_date
                    queryset = queryset.filter(last_update__date__lte=end_date)

            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD format."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        today = get_today_local()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        year_ago = today - timedelta(days=365)

        # Basic counts with corrected status references
        basic_stats = {
            "total": queryset.count(),
            "completed": queryset.filter(status=TaskStatus.COMPLETED).count(),
            "in_progress": queryset.filter(status=TaskStatus.ON_GOING).count(),
            "pending": queryset.filter(status=TaskStatus.PENDING).count(),
            "for_checking": queryset.filter(status=TaskStatus.FOR_CHECKING).count(),
            "for_revision": queryset.filter(status=TaskStatus.FOR_REVISION).count(),
            "not_started": queryset.filter(status=TaskStatus.NOT_YET_STARTED).count(),
            "cancelled": queryset.filter(status=TaskStatus.CANCELLED).count(),
        }

        # Priority distribution
        priority_stats = {
            "high_priority": queryset.filter(priority=TaskPriority.HIGH).count(),
            "medium_priority": queryset.filter(priority=TaskPriority.MEDIUM).count(),
            "low_priority": queryset.filter(priority=TaskPriority.LOW).count(),
        }

        # Category distribution with display names
        category_stats = (
            queryset.values("category")
            .annotate(count=Count("category"))
            .order_by("-count")
        )
        category_distribution = {}
        for item in category_stats:
            # Get display name for category
            category_display = dict(TaskCategory.choices).get(
                item["category"], item["category"]
            )
            category_distribution[item["category"]] = {
                "count": item["count"],
                "display_name": category_display,
            }

        # Time-based analysis
        overdue_tasks = queryset.filter(
            deadline__lt=today,
            status__in=[
                TaskStatus.NOT_YET_STARTED,
                TaskStatus.ON_GOING,
                TaskStatus.PENDING,
            ],
        ).count()

        due_today = queryset.filter(
            deadline=today,
            status__in=[
                TaskStatus.NOT_YET_STARTED,
                TaskStatus.ON_GOING,
                TaskStatus.PENDING,
            ],
        ).count()

        due_this_week = queryset.filter(
            deadline__range=[today, today + timedelta(days=7)],
            status__in=[
                TaskStatus.NOT_YET_STARTED,
                TaskStatus.ON_GOING,
                TaskStatus.PENDING,
            ],
        ).count()

        due_this_month = queryset.filter(
            deadline__range=[today, today + timedelta(days=30)],
            status__in=[
                TaskStatus.NOT_YET_STARTED,
                TaskStatus.ON_GOING,
                TaskStatus.PENDING,
            ],
        ).count()

        # Recent activity metrics - using last_update as proxy for creation
        created_last_week = queryset.filter(last_update__gte=week_ago).count()
        completed_last_week = queryset.filter(
            completion_date__gte=week_ago, status=TaskStatus.COMPLETED
        ).count()
        completed_last_month = queryset.filter(
            completion_date__gte=month_ago, status=TaskStatus.COMPLETED
        ).count()

        # Client distribution (Top 10 clients)
        client_stats = (
            queryset.values("client__name", "client__id", "client__status")
            .annotate(
                total_tasks=Count("id"),
                completed_tasks=Count(
                    Case(
                        When(status=TaskStatus.COMPLETED, then=1),
                        output_field=IntegerField(),
                    )
                ),
                overdue_tasks=Count(
                    Case(
                        When(
                            Q(deadline__lt=today)
                            & Q(
                                status__in=[
                                    TaskStatus.NOT_YET_STARTED,
                                    TaskStatus.ON_GOING,
                                    TaskStatus.PENDING,
                                ]
                            ),
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
                pending_tasks=Count(
                    Case(
                        When(
                            status__in=[
                                TaskStatus.NOT_YET_STARTED,
                                TaskStatus.ON_GOING,
                                TaskStatus.PENDING,
                            ],
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
            )
            .order_by("-total_tasks")[:10]
        )

        # Add completion rate for clients
        for client in client_stats:
            if client["total_tasks"] > 0:
                client["completion_rate"] = round(
                    (client["completed_tasks"] / client["total_tasks"]) * 100, 2
                )
            else:
                client["completion_rate"] = 0

        # User performance with enhanced metrics
        user_stats = (
            queryset.values(
                "assigned_to__first_name",
                "assigned_to__last_name",
                "assigned_to__id",
                "assigned_to__role",
            )
            .annotate(
                total_tasks=Count("id"),
                completed_tasks=Count(
                    Case(
                        When(status=TaskStatus.COMPLETED, then=1),
                        output_field=IntegerField(),
                    )
                ),
                overdue_tasks=Count(
                    Case(
                        When(
                            Q(deadline__lt=today)
                            & Q(
                                status__in=[
                                    TaskStatus.NOT_YET_STARTED,
                                    TaskStatus.ON_GOING,
                                    TaskStatus.PENDING,
                                ]
                            ),
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
                pending_tasks=Count(
                    Case(
                        When(
                            status__in=[
                                TaskStatus.NOT_YET_STARTED,
                                TaskStatus.ON_GOING,
                                TaskStatus.PENDING,
                            ],
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
                high_priority_tasks=Count(
                    Case(
                        When(priority=TaskPriority.HIGH, then=1),
                        output_field=IntegerField(),
                    )
                ),
            )
            .order_by("-total_tasks")
        )

        # Enhanced user stats with calculated metrics
        for user in user_stats:
            total = user["total_tasks"]
            if total > 0:
                user["completion_rate"] = round(
                    (user["completed_tasks"] / total) * 100, 2
                )
                user["overdue_rate"] = round((user["overdue_tasks"] / total) * 100, 2)
            else:
                user["completion_rate"] = 0
                user["overdue_rate"] = 0
            user["fullname"] = (
                f"{user['assigned_to__first_name']} {user['assigned_to__last_name']}"
            )
            user["is_admin"] = user["assigned_to__role"] == "admin"

        # Weekly completion trend (last 8 weeks for better chart visualization)
        weekly_trends = []
        for i in range(8):
            week_start = today - timedelta(days=(i * 7) + today.weekday())
            week_end = week_start + timedelta(days=6)

            week_completed = queryset.filter(
                completion_date__range=[week_start, week_end],
                status=TaskStatus.COMPLETED,
            ).count()

            week_created = queryset.filter(
                last_update__range=[week_start, week_end]
            ).count()

            weekly_trends.append(
                {
                    "week_start": week_start.strftime("%Y-%m-%d"),
                    "week_label": f"Week of {week_start.strftime('%b %d')}",
                    "completed": week_completed,
                    "created": week_created,
                }
            )

        weekly_trends.reverse()  # Show chronologically

        # Approval workflow statistics
        approval_stats = {
            "tasks_requiring_approval": queryset.filter(requires_approval=True).count(),
            "tasks_in_approval": queryset.filter(
                requires_approval=True, status=TaskStatus.FOR_CHECKING
            ).count(),
            "tasks_approved": queryset.filter(
                requires_approval=True, status=TaskStatus.COMPLETED
            ).count(),
            "pending_my_approval": 0,  # Will be calculated for current user
        }

        # Calculate pending approvals for current user if admin
        if request.user.is_admin:
            from core.models import TaskApproval

            approval_stats["pending_my_approval"] = TaskApproval.objects.filter(
                approver=request.user, action="pending"
            ).count()

        # Tax-specific analytics
        tax_stats = {}
        tax_tasks = queryset.filter(category=TaskCategory.TAX_CASE)
        if tax_tasks.exists():
            total_payable = tax_tasks.aggregate(total=Sum("tax_payable"))["total"] or 0

            tax_stats = {
                "total_tax_cases": tax_tasks.count(),
                "completed_tax_cases": tax_tasks.filter(
                    status=TaskStatus.COMPLETED
                ).count(),
                "tax_payable_total": float(total_payable),
                "average_tax_payable": (
                    float(total_payable / tax_tasks.count())
                    if tax_tasks.count() > 0
                    else 0
                ),
                "by_tax_category": [
                    {
                        "category": item["tax_category"],
                        "display_name": (
                            dict(TaxCaseCategory.choices).get(
                                item["tax_category"], item["tax_category"]
                            )
                            if item["tax_category"]
                            else "Not Specified"
                        ),
                        "count": item["count"],
                    }
                    for item in tax_tasks.values("tax_category").annotate(
                        count=Count("tax_category")
                    )
                    if item["tax_category"]
                ],
                "by_tax_type": [
                    {
                        "type": item["tax_type"],
                        "display_name": (
                            dict(TypeOfTaxCase.choices).get(
                                item["tax_type"], item["tax_type"]
                            )
                            if item["tax_type"]
                            else "Not Specified"
                        ),
                        "count": item["count"],
                    }
                    for item in tax_tasks.values("tax_type").annotate(
                        count=Count("tax_type")
                    )
                    if item["tax_type"]
                ],
            }

        # Productivity and performance metrics
        completed_tasks = queryset.filter(status=TaskStatus.COMPLETED)
        total_tasks = queryset.count()

        productivity_stats = {
            "overall_completion_rate": round(
                (
                    (basic_stats["completed"] / total_tasks * 100)
                    if total_tasks > 0
                    else 0
                ),
                2,
            ),
            "on_time_completion_rate": 0,
            "average_completion_days": 0,
            "tasks_completed_early": 0,
            "tasks_completed_late": 0,
            "workload_balance_score": 0,  # How evenly distributed tasks are among users
        }

        if completed_tasks.exists():
            # On-time completion rate
            on_time_completed = completed_tasks.filter(
                completion_date__lte=F("deadline")
            ).count()
            late_completed = completed_tasks.filter(
                completion_date__gt=F("deadline")
            ).count()
            early_completed = completed_tasks.filter(
                completion_date__lt=F("deadline")
            ).count()

            productivity_stats.update(
                {
                    "on_time_completion_rate": round(
                        (on_time_completed / completed_tasks.count()) * 100, 2
                    ),
                    "tasks_completed_early": early_completed,
                    "tasks_completed_late": late_completed,
                }
            )

            # Average completion time calculation - using engagement_date if available
            completed_with_dates = completed_tasks.filter(
                completion_date__isnull=False, engagement_date__isnull=False
            )

            if completed_with_dates.exists():
                total_days = sum(
                    [
                        (task.completion_date - task.engagement_date).days
                        for task in completed_with_dates
                        if task.completion_date and task.engagement_date
                    ]
                )
                if completed_with_dates.count() > 0:
                    productivity_stats["average_completion_days"] = round(
                        total_days / completed_with_dates.count(), 1
                    )

        # Calculate workload balance (standard deviation of task distribution)
        if user_stats:
            task_counts = [user["total_tasks"] for user in user_stats]
            if len(task_counts) > 1:
                mean_tasks = sum(task_counts) / len(task_counts)
                variance = sum((x - mean_tasks) ** 2 for x in task_counts) / len(
                    task_counts
                )
                std_deviation = variance**0.5
                # Convert to a 0-100 scale where 100 is perfectly balanced
                productivity_stats["workload_balance_score"] = (
                    max(0, round(100 - (std_deviation / mean_tasks * 100), 2))
                    if mean_tasks > 0
                    else 100
                )

        # System health metrics
        system_health = {
            "active_clients": queryset.values("client")
            .filter(client__status="active")
            .distinct()
            .count(),
            "total_clients": queryset.values("client").distinct().count(),
            "average_tasks_per_client": (
                round(total_tasks / queryset.values("client").distinct().count(), 2)
                if queryset.values("client").distinct().count() > 0
                else 0
            ),
            "critical_overdue": queryset.filter(
                deadline__lt=today - timedelta(days=7),
                priority=TaskPriority.HIGH,
                status__in=[
                    TaskStatus.NOT_YET_STARTED,
                    TaskStatus.ON_GOING,
                    TaskStatus.PENDING,
                ],
            ).count(),
            "system_load_indicator": "low",  # Will be calculated based on various factors
        }

        # Calculate system load indicator
        overdue_percentage = (
            (overdue_tasks / total_tasks * 100) if total_tasks > 0 else 0
        )
        if overdue_percentage > 20 or system_health["critical_overdue"] > 5:
            system_health["system_load_indicator"] = "high"
        elif overdue_percentage > 10 or system_health["critical_overdue"] > 2:
            system_health["system_load_indicator"] = "medium"

        # Quick actions data (useful for dashboard widgets)
        quick_actions = {
            "tasks_need_attention": queryset.filter(
                Q(deadline__lte=today + timedelta(days=3))
                & Q(
                    status__in=[
                        TaskStatus.NOT_YET_STARTED,
                        TaskStatus.ON_GOING,
                        TaskStatus.PENDING,
                    ]
                )
            ).count(),
            "high_priority_pending": queryset.filter(
                priority=TaskPriority.HIGH,
                status__in=[
                    TaskStatus.NOT_YET_STARTED,
                    TaskStatus.ON_GOING,
                    TaskStatus.PENDING,
                ],
            ).count(),
            "recent_completions": completed_last_week,
            "new_tasks_this_week": created_last_week,
        }

        # Compile comprehensive dashboard statistics
        dashboard_stats = {
            "summary": {
                **basic_stats,
                **priority_stats,
                "overdue": overdue_tasks,
                "due_today": due_today,
                "due_this_week": due_this_week,
                "due_this_month": due_this_month,
            },
            "charts_data": {
                "category_distribution": category_distribution,
                "weekly_trends": weekly_trends,
                "status_breakdown": {
                    "completed": basic_stats["completed"],
                    "in_progress": basic_stats["in_progress"],
                    "pending": basic_stats["pending"],
                    "not_started": basic_stats["not_started"],
                    "for_checking": basic_stats["for_checking"],
                    "for_revision": basic_stats["for_revision"],
                },
                "priority_breakdown": priority_stats,
            },
            "performance_metrics": productivity_stats,
            "team_analytics": {
                "user_performance": list(user_stats),
                "workload_distribution": [
                    {
                        "user": user["fullname"],
                        "tasks": user["total_tasks"],
                        "completion_rate": user["completion_rate"],
                    }
                    for user in user_stats[:5]  # Top 5 for widget display
                ],
            },
            "client_insights": {
                "top_clients": list(client_stats)[:5],  # Top 5 for dashboard
                "all_clients": list(client_stats),  # Full list for detailed view
            },
            "business_intelligence": {
                "approval_workflow": approval_stats,
                "tax_analysis": tax_stats,
                "system_health": system_health,
                "recent_activity": {
                    "completed_last_week": completed_last_week,
                    "completed_last_month": completed_last_month,
                    "created_last_week": created_last_week,
                },
            },
            "quick_actions": quick_actions,
            "metadata": {
                "generated_at": today.isoformat(),
                "user_role": request.user.role,
                "is_admin": request.user.is_admin,
                "data_scope": (
                    "all_tasks" if request.user.is_admin else "assigned_tasks"
                ),
                "filters_applied": filters_applied,
            },
        }

        return Response(dashboard_stats)

    @action(detail=False, methods=["post"], url_path="export-statistics")
    def export_statistics(self, request):
        """Export comprehensive task statistics to CSV or Excel format"""
        import csv
        from datetime import datetime, timedelta
        from io import BytesIO

        from django.db.models import Avg, Case, Count, F, IntegerField, Q, Sum, When
        from django.db.models.functions import Extract, TruncMonth
        from django.http import HttpResponse
        from openpyxl import Workbook

        from core.choices import (
            TaskCategory,
            TaskPriority,
            TaskStatus,
            TaxCaseCategory,
            TypeOfTaxCase,
        )

        # Get format parameter from request data (default to csv)
        export_format = request.data.get("format", "csv")
        if isinstance(export_format, str):
            export_format = export_format.lower().strip()
        else:
            export_format = "csv"

        if not export_format or export_format not in ["csv", "excel"]:
            return Response(
                {"error": "Invalid format. Use 'csv' or 'excel'"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Reuse the same queryset and filtering logic as statistics method
        queryset = self.get_queryset()

        # Parse and validate date range filter parameters
        start_date_param = request.data.get("start_date")
        end_date_param = request.data.get("end_date")

        start_date = None
        end_date = None

        if start_date_param or end_date_param:
            try:
                if start_date_param:
                    start_date = datetime.strptime(start_date_param, "%Y-%m-%d").date()
                if end_date_param:
                    end_date = datetime.strptime(end_date_param, "%Y-%m-%d").date()
                if start_date and end_date and start_date > end_date:
                    return Response(
                        {"error": "start_date cannot be after end_date"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if start_date and end_date:
                    queryset = queryset.filter(deadline__range=[start_date, end_date])
                elif start_date:
                    queryset = queryset.filter(deadline__gte=start_date)
                elif end_date:
                    queryset = queryset.filter(deadline__lte=end_date)
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD format."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        today = get_today_local()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Get basic statistics
        basic_stats = {
            "total": queryset.count(),
            "completed": queryset.filter(status=TaskStatus.COMPLETED).count(),
            "in_progress": queryset.filter(status=TaskStatus.ON_GOING).count(),
            "pending": queryset.filter(status=TaskStatus.PENDING).count(),
            "for_checking": queryset.filter(status=TaskStatus.FOR_CHECKING).count(),
            "for_revision": queryset.filter(status=TaskStatus.FOR_REVISION).count(),
            "not_started": queryset.filter(status=TaskStatus.NOT_YET_STARTED).count(),
            "cancelled": queryset.filter(status=TaskStatus.CANCELLED).count(),
        }

        # Get user performance data
        user_stats = (
            queryset.values(
                "assigned_to__first_name",
                "assigned_to__last_name",
                "assigned_to__id",
                "assigned_to__role",
            )
            .annotate(
                total_tasks=Count("id"),
                completed_tasks=Count(
                    Case(
                        When(status=TaskStatus.COMPLETED, then=1),
                        output_field=IntegerField(),
                    )
                ),
                overdue_tasks=Count(
                    Case(
                        When(
                            Q(deadline__lt=today)
                            & Q(
                                status__in=[
                                    TaskStatus.NOT_YET_STARTED,
                                    TaskStatus.ON_GOING,
                                    TaskStatus.PENDING,
                                ]
                            ),
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
                pending_tasks=Count(
                    Case(
                        When(
                            status__in=[
                                TaskStatus.NOT_YET_STARTED,
                                TaskStatus.ON_GOING,
                                TaskStatus.PENDING,
                            ],
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
            )
            .order_by("-total_tasks")
        )

        # Add completion rate for users
        for user in user_stats:
            total = user["total_tasks"]
            if total > 0:
                user["completion_rate"] = round(
                    (user["completed_tasks"] / total) * 100, 2
                )
                user["overdue_rate"] = round((user["overdue_tasks"] / total) * 100, 2)
            else:
                user["completion_rate"] = 0
                user["overdue_rate"] = 0
            user["fullname"] = (
                f"{user['assigned_to__first_name']} {user['assigned_to__last_name']}"
            )

        # Get client statistics
        client_stats = (
            queryset.values("client__name", "client__id", "client__status")
            .annotate(
                total_tasks=Count("id"),
                completed_tasks=Count(
                    Case(
                        When(status=TaskStatus.COMPLETED, then=1),
                        output_field=IntegerField(),
                    )
                ),
                overdue_tasks=Count(
                    Case(
                        When(
                            Q(deadline__lt=today)
                            & Q(
                                status__in=[
                                    TaskStatus.NOT_YET_STARTED,
                                    TaskStatus.ON_GOING,
                                    TaskStatus.PENDING,
                                ]
                            ),
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
                pending_tasks=Count(
                    Case(
                        When(
                            status__in=[
                                TaskStatus.NOT_YET_STARTED,
                                TaskStatus.ON_GOING,
                                TaskStatus.PENDING,
                            ],
                            then=1,
                        ),
                        output_field=IntegerField(),
                    )
                ),
            )
            .order_by("-total_tasks")[:50]  # Limit to top 50 for export
        )

        # Add completion rate for clients
        for client in client_stats:
            if client["total_tasks"] > 0:
                client["completion_rate"] = round(
                    (client["completed_tasks"] / client["total_tasks"]) * 100, 2
                )
            else:
                client["completion_rate"] = 0

        # Prepare data for export
        if export_format == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = (
                'attachment; filename="task_statistics.csv"'
            )
            writer = csv.writer(response)

            # Write summary statistics
            writer.writerow(["Task Statistics Export"])
            writer.writerow(["Generated on", today.strftime("%Y-%m-%d %H:%M:%S")])

            # Add date range information if provided
            if start_date or end_date:
                if start_date and end_date:
                    writer.writerow(
                        [
                            "Date Range",
                            f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                        ]
                    )
                elif start_date:
                    writer.writerow(["Start Date", start_date.strftime("%Y-%m-%d")])
                elif end_date:
                    writer.writerow(["End Date", end_date.strftime("%Y-%m-%d")])
                writer.writerow([])

            # Basic stats
            writer.writerow(["Summary Statistics"])
            writer.writerow(["Metric", "Value"])
            for key, value in basic_stats.items():
                writer.writerow([key.replace("_", " ").title(), value])
            writer.writerow([])

            # User performance
            writer.writerow(["User Performance"])
            writer.writerow(
                [
                    "User",
                    "Total Tasks",
                    "Completed",
                    "Pending",
                    "Overdue",
                    "Completion Rate (%)",
                    "Overdue Rate (%)",
                ]
            )
            for user in user_stats:
                writer.writerow(
                    [
                        user["fullname"],
                        user["total_tasks"],
                        user["completed_tasks"],
                        user["pending_tasks"],
                        user["overdue_tasks"],
                        user["completion_rate"],
                        user["overdue_rate"],
                    ]
                )
            writer.writerow([])

            # Client statistics
            writer.writerow(["Client Statistics"])
            writer.writerow(
                [
                    "Client",
                    "Total Tasks",
                    "Completed",
                    "Pending",
                    "Overdue",
                    "Completion Rate (%)",
                ]
            )
            for client in client_stats:
                writer.writerow(
                    [
                        client["client__name"],
                        client["total_tasks"],
                        client["completed_tasks"],
                        client["pending_tasks"],
                        client["overdue_tasks"],
                        client["completion_rate"],
                    ]
                )

        else:  # Excel format
            wb = Workbook()
            ws_summary = wb.active
            if ws_summary:
                ws_summary.title = "Summary"
            ws_users = wb.create_sheet("User Performance")
            ws_clients = wb.create_sheet("Client Statistics")

            # Summary sheet
            if ws_summary:
                ws_summary["A1"] = "Task Statistics Export"
                ws_summary["A2"] = (
                    f"Generated on: {today.strftime('%Y-%m-%d %H:%M:%S')}"
                )

                # Add date range information if provided
                row_num = 4
                if start_date or end_date:
                    if start_date and end_date:
                        ws_summary[f"A{row_num}"] = (
                            f"Date Range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
                        )
                    elif start_date:
                        ws_summary[f"A{row_num}"] = (
                            f"Start Date: {start_date.strftime('%Y-%m-%d')}"
                        )
                    elif end_date:
                        ws_summary[f"A{row_num}"] = (
                            f"End Date: {end_date.strftime('%Y-%m-%d')}"
                        )
                    row_num += 1

                ws_summary[f"A{row_num}"] = "Summary Statistics"
                ws_summary[f"A{row_num + 1}"] = "Metric"
                ws_summary[f"B{row_num + 1}"] = "Value"

                row = row_num + 2
                for key, value in basic_stats.items():
                    ws_summary[f"A{row}"] = key.replace("_", " ").title()
                    ws_summary[f"B{row}"] = value
                    row += 1

            # User performance sheet
            ws_users["A1"] = "User Performance"
            headers = [
                "User",
                "Total Tasks",
                "Completed",
                "Pending",
                "Overdue",
                "Completion Rate (%)",
                "Overdue Rate (%)",
            ]
            for col, header in enumerate(headers, 1):
                ws_users.cell(row=1, column=col, value=header)

            for row, user in enumerate(user_stats, 2):
                ws_users.cell(row=row, column=1, value=user["fullname"])
                ws_users.cell(row=row, column=2, value=user["total_tasks"])
                ws_users.cell(row=row, column=3, value=user["completed_tasks"])
                ws_users.cell(row=row, column=4, value=user["pending_tasks"])
                ws_users.cell(row=row, column=5, value=user["overdue_tasks"])
                ws_users.cell(row=row, column=6, value=user["completion_rate"])
                ws_users.cell(row=row, column=7, value=user["overdue_rate"])

            # Client statistics sheet
            ws_clients["A1"] = "Client Statistics"
            headers = [
                "Client",
                "Total Tasks",
                "Completed",
                "Pending",
                "Overdue",
                "Completion Rate (%)",
            ]
            for col, header in enumerate(headers, 1):
                ws_clients.cell(row=1, column=col, value=header)

            for row, client in enumerate(client_stats, 2):
                ws_clients.cell(row=row, column=1, value=client["client__name"])
                ws_clients.cell(row=row, column=2, value=client["total_tasks"])
                ws_clients.cell(row=row, column=3, value=client["completed_tasks"])
                ws_clients.cell(row=row, column=4, value=client["pending_tasks"])
                ws_clients.cell(row=row, column=5, value=client["overdue_tasks"])
                ws_clients.cell(row=row, column=6, value=client["completion_rate"])

            # Save to response
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="task_statistics.xlsx"'
            )

        return response

    @action(detail=True, methods=["POST"], url_path="update-deadline")
    def update_deadline(self, request, pk=None):
        task = self.get_object()
        try:
            updated_status = request.data.get("status")
            updated_remarks = request.data.get("remarks")

            # Update task fields
            if updated_remarks:
                task.remarks = updated_remarks

            # Use add_status_update to handle both status and remarks updates
            # The method now has built-in logic to prevent unnecessary status history entries
            task.add_status_update(
                new_status=updated_status,
                remarks=updated_remarks,
                changed_by=request.user,
            )

            serializer = TaskListSerializer(task)
            return Response(
                data=serializer.data,
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                data={"message": f"Something went wrong. {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["POST"], url_path="initiate-approval")
    def initiate_approval(self, request, pk=None):
        """Initiate approval workflow for a task"""
        task = self.get_object()

        # Check if user has permission to initiate approval
        if not request.user.is_admin and task.assigned_to != request.user:
            return Response(
                {
                    "error": "You don't have permission to initiate approval for this task."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check if task is already in approval workflow
        if task.requires_approval:
            return Response(
                {"error": "This task is already in approval workflow."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate request data
        serializer = InitiateApprovalSerializer(data=request.data)
        if serializer.is_valid():
            approver_ids = serializer.validated_data["approvers"]
            approvers = User.objects.filter(id__in=approver_ids, role="admin")

            try:
                initiate_task_approval(task, list(approvers), request.user)
                return Response(
                    {"message": "Approval workflow initiated successfully."},
                    status=status.HTTP_200_OK,
                )
            except Exception as e:
                return Response(
                    {"error": f"Failed to initiate approval workflow: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["POST"], url_path="process-approval")
    def process_approval(self, request, pk=None):
        """Process an approval decision for a task"""
        task = self.get_object()

        # Check if user is an admin
        if not request.user.is_admin:
            return Response(
                {"error": "Only admin users can process approvals."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check if task is in approval workflow
        if not task.requires_approval:
            return Response(
                {"error": "This task is not in approval workflow."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if user is the current approver
        current_approval = TaskApproval.objects.filter(
            task=task,
            approver=request.user,
            step_number=task.current_approval_step,
            action="pending",
        ).first()

        if not current_approval:
            return Response(
                {"error": "You are not the current approver for this task."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Validate request data
        serializer = ProcessApprovalSerializer(data=request.data)
        if serializer.is_valid():
            action = serializer.validated_data["action"]
            comments = serializer.validated_data.get("comments", "")
            next_approver = serializer.validated_data.get("next_approver")

            try:
                process_task_approval(
                    task, request.user, action, comments, next_approver
                )
                return Response(
                    {"message": f"Task {action} successfully."},
                    status=status.HTTP_200_OK,
                )
            except Exception as e:
                return Response(
                    {"error": f"Failed to process approval: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["GET"], url_path="status-history")
    def status_history(self, request, pk=None):
        """Get status history for a specific task"""
        task = self.get_object()

        # Get all status history records for this task, ordered by creation date
        status_history = TaskStatusHistory.objects.filter(task=task).order_by(
            "-created_at"
        )

        # Serialize the status history records
        serializer = TaskStatusHistorySerializer(status_history, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["GET"], url_path="task-approvals")
    def task_approvals(self, request, pk=None):
        """Get approval records for a specific task"""
        task = self.get_object()

        # Get all approval records for this task, ordered by step number
        approvals = TaskApproval.objects.filter(task=task).order_by("step_number")

        # Serialize the approval records
        serializer = TaskApprovalSerializer(approvals, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["GET"], url_path="pending-approvals")
    def pending_approvals(self, request):
        """Get tasks pending current user's approval"""
        if not request.user.is_admin:
            return Response(
                {"error": "Only admin users can view pending approvals."},
                status=status.HTTP_403_FORBIDDEN,
            )

        pending_approvals = TaskApproval.objects.filter(
            approver=request.user, action="pending"
        ).select_related("task", "task__client", "task__assigned_to")

        tasks = [approval.task for approval in pending_approvals]

        # Optimize queries for the TaskListSerializer
        task_ids = [task.id for task in tasks]
        optimized_tasks = (
            Task.objects.filter(id__in=task_ids)
            .select_related("assigned_to", "client")
            .prefetch_related(
                "approvals__approver",
                "approvals__next_approver",
                "status_history_records__changed_by",
            )
        )

        return Response(
            TaskListSerializer(optimized_tasks, many=True).data,
            status=status.HTTP_200_OK,
        )

    def destroy(self, request, *args, **kwargs):
        """Override destroy to send notification before deleting task"""
        instance = self.get_object()

        # Send notification to assigned user before deletion
        if instance.assigned_to and instance.assigned_to != request.user:
            create_notifications(
                recipient=instance.assigned_to,
                title="Task Deleted",
                message=f"The task '{instance.description}' has been deleted.",
                link="/my-deadlines",
            )

        # Log the deletion
        create_log(request.user, f"Deleted task: {instance}")

        # Perform the deletion
        return super().destroy(request, *args, **kwargs)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.select_related("created_by")
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrStaff]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    search_fields = ["name", "contact_person", "email", "address", "phone"]
    ordering_fields = ["name", "created_at"]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Non-staff users only see clients they created
        if not self.request.user.is_admin:
            queryset = queryset.filter(created_by=self.request.user).distinct()

        return queryset

    @action(detail=False, methods=["get"], url_path="birthdays")
    def get_birthdays(self, request):
        today = get_today_local()
        current_month = today.month
        current_day = today.day

        # Get all clients (we'll filter by month/day comparison)
        all_clients = self.get_queryset()

        # Categorize by comparing only month and day (ignore year)
        birthdays_today = []
        upcoming_birthdays = []
        past_birthdays = []

        for client in all_clients:
            # Skip clients without birth date
            if not client.date_of_birth:
                continue

            # Calculate days until birthday (handling year wrap-around)
            try:
                # Create this year's birthday date
                this_year_birthday = client.date_of_birth.replace(year=today.year)

                # If birthday has passed this year, calculate for next year
                if this_year_birthday < today:
                    next_birthday = client.date_of_birth.replace(year=today.year + 1)
                else:
                    next_birthday = this_year_birthday

                days_until_birthday = (next_birthday - today).days

                # Only include birthdays within 7-day window
                if days_until_birthday == 0:
                    # Today's birthday
                    birthdays_today.append(client)
                elif 1 <= days_until_birthday <= 7:
                    # Upcoming: tomorrow to 7 days from today
                    upcoming_birthdays.append(client)
                elif -7 <= days_until_birthday <= -1:
                    # Past: 7 days ago to yesterday
                    past_birthdays.append(client)
                # Exclude birthdays beyond 7 days in either direction

            except ValueError:
                # Handle invalid dates (e.g., Feb 30)
                pass

        # Sort each category by date (month and day)
        def sort_by_birth_date(client):
            return client.date_of_birth.month * 100 + client.date_of_birth.day

        birthdays_today.sort(key=sort_by_birth_date)
        upcoming_birthdays.sort(key=sort_by_birth_date)
        past_birthdays.sort(key=sort_by_birth_date)

        return Response(
            {
                "today": ClientBirthdaySerializer(birthdays_today, many=True).data,
                "upcoming": ClientBirthdaySerializer(
                    upcoming_birthdays, many=True
                ).data,
                "past": ClientBirthdaySerializer(past_birthdays, many=True).data,
            },
            status=status.HTTP_200_OK,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
            create_log(self.request.user, f"Deleted client: {instance}.")
        except RestrictedError as e:
            return Response(
                {
                    "detail": "Cannot delete this object because it is referenced by other records.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        create_log(self.request.user, f"Created client: {instance}.")

    def perform_update(self, serializer):
        instance = serializer.save()
        create_log(
            self.request.user,
            f"Updated client: {instance}. Status: {'Active' if instance.is_active else 'Inactive'}.",
        )


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.select_related("recipient")
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
    ]

    filterset_fields = ["recipient", "is_read"]

    @action(detail=True, methods=["post"], url_path="mark-as-read")
    def mark_as_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_read()

        # Log the action
        create_log(
            request.user,
            f"Marked notification '{notification.title}' as read",
        )

        return Response(status=status.HTTP_200_OK)


class AppLogViewSet(viewsets.ModelViewSet):
    queryset = AppLog.objects.select_related("user")
    serializer_class = AppLogSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrStaff]
    filter_backends = [
        DjangoFilterBackend,
    ]
    filterset_fields = ["user"]

    @action(detail=False, methods=["get"], url_path="users")
    def get_user_choices(self, request):
        users = User.objects.exclude(logs__isnull=True)
        serializer = UserMiniSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ClientDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing ClientDocument records

    Provides CRUD operations for client documents with file upload capabilities.
    """

    queryset = ClientDocument.objects.select_related("client", "uploaded_by")
    serializer_class = ClientDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    # Filtering options
    filterset_fields = {
        "client": ["exact"],
        "uploaded_by": ["exact"],
        "uploaded_at": ["gte", "lte", "exact"],
    }

    # Search fields
    search_fields = [
        "title",
        "description",
        "client__name",
        "uploaded_by__first_name",
        "uploaded_by__last_name",
    ]

    # Ordering options
    ordering_fields = [
        "title",
        "uploaded_at",
        "updated_at",
        "client__name",
    ]
    ordering = ["-uploaded_at"]

    def get_queryset(self):
        """
        Filter queryset based on user permissions.
        Admin users see all documents, non-admin users only see documents for clients they created.
        By default, excludes soft-deleted documents.
        """
        queryset = ClientDocument.objects.select_related("client", "uploaded_by")

        if not self.request.user.is_authenticated:
            return queryset.none()

        # Filter out soft-deleted documents by default
        queryset = queryset.filter(is_deleted=False)

        if self.request.user.is_admin:
            return queryset

        # Non-admin users can only see documents for clients they created
        return queryset.filter(client__created_by=self.request.user)

    def perform_create(self, serializer):
        """Create document and log the action"""
        instance = serializer.save()
        create_log(
            self.request.user,
            f"Uploaded document '{instance.title}' for client {instance.client.name}.",
        )

    def perform_update(self, serializer):
        """Update document and log the action"""
        instance = serializer.save()
        create_log(
            self.request.user,
            f"Updated document '{instance.title}' for client {instance.client.name}.",
        )

    def perform_destroy(self, instance):
        """Soft delete document and log the action"""
        client_name = instance.client.name
        document_title = instance.title

        # Perform soft delete
        success = instance.soft_delete()

        if success:
            create_log(
                self.request.user,
                f"Soft deleted document '{document_title}' for client {client_name}.",
            )
        else:
            create_log(
                self.request.user,
                f"Soft deleted document '{document_title}' for client {client_name} (file move failed).",
            )

    @action(detail=False, methods=["get"], url_path="by-client")
    def get_documents_by_client(self, request):
        """Get all documents for a specific client"""
        client_id = request.query_params.get("client_id")
        if not client_id:
            return Response(
                {"error": "client_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return Response(
                {"error": "Client not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Check permissions
        if not request.user.is_admin and client.created_by != request.user:
            return Response(
                {
                    "error": "You don't have permission to view documents for this client"
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        documents = self.get_queryset().filter(client=client)
        serializer = self.get_serializer(documents, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="download")
    def download_document(self, request, pk=None):
        """Download a specific document"""
        document = self.get_object()

        # Check permissions
        if not request.user.is_admin and document.client.created_by != request.user:
            return Response(
                {"error": "You don't have permission to download this document"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check if file exists in storage
        if not document.document_file or not document.document_file.storage.exists(
            document.document_file.name
        ):
            return Response(
                {
                    "error": "File not found in storage. The file may have been moved or deleted."
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            file_handle = document.document_file.open()
            response = FileResponse(
                file_handle,
                content_type="application/octet-stream",
            )
            filename = (
                document.document_file.name.split("/")[-1]
                if document.document_file.name
                else "download"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Log the download
            create_log(
                request.user,
                f"Downloaded document '{document.title}' for client {document.client.name}",
            )

            return response
        except Exception as e:
            return Response(
                {"error": f"Error downloading file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="deleted")
    def get_deleted_documents(self, request):
        """Get all soft-deleted documents (admin only)"""
        if not request.user.is_admin:
            return Response(
                {"error": "Only admin users can view deleted documents"},
                status=status.HTTP_403_FORBIDDEN,
            )

        deleted_docs = ClientDocument.objects.filter(is_deleted=True).select_related(
            "client", "uploaded_by"
        )

        serializer = self.get_serializer(deleted_docs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="restore")
    def restore_document(self, request, pk=None):
        """Restore a soft-deleted document"""
        document = self.get_object()

        # Check if document is actually deleted
        if not document.is_deleted:
            return Response(
                {"error": "Document is not deleted"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check permissions
        if not request.user.is_admin and document.client.created_by != request.user:
            return Response(
                {"error": "You don't have permission to restore this document"},
                status=status.HTTP_403_FORBIDDEN,
            )

        success = document.restore()

        if success:
            create_log(
                request.user,
                f"Restored document '{document.title}' for client {document.client.name}.",
            )
            serializer = self.get_serializer(document)
            return Response(serializer.data)
        else:
            return Response(
                {"error": "Failed to restore document"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["delete"], url_path="hard-delete")
    def hard_delete_document(self, request, pk=None):
        """Permanently delete a document and its file (admin only)"""
        if not request.user.is_admin:
            return Response(
                {"error": "Only admin users can permanently delete documents"},
                status=status.HTTP_403_FORBIDDEN,
            )

        document = self.get_object()

        # Check permissions
        if not request.user.is_admin and document.client.created_by != request.user:
            return Response(
                {"error": "You don't have permission to delete this document"},
                status=status.HTTP_403_FORBIDDEN,
            )

        client_name = document.client.name
        document_title = document.title

        # Perform hard delete
        document.hard_delete()

        create_log(
            request.user,
            f"Permanently deleted document '{document_title}' for client {client_name}.",
        )

        return Response(status=status.HTTP_204_NO_CONTENT)
